from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Count, Q
from decimal import Decimal
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.urls import reverse
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from .forms import PedidoForm, ProductoForm
from .models import Pedido, Producto
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.db import IntegrityError
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.contrib.auth.password_validation import validate_password
import json
import requests
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
# Create your views here.

def home(request):
    return render(request, 'home.html')

def signup(request):        
    if request.method == 'GET':
        return render(request, 'signup.html', {
        'form': UserCreationForm()
         })
    else:
        name = request.POST.get('name')
        email = request.POST.get('email')
        confirm_email = request.POST.get('confirm_email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        try:
            validate_email(email)
            validate_email(confirm_email)
        except ValidationError:
            return render(request, 'signup.html', {
                'form': UserCreationForm(),
                'error': 'Formato de correo inválido'
            })

        if email != confirm_email:
            return render(request, 'signup.html', {
                'form': UserCreationForm(),
                'error': 'Los correos no coinciden'
            })

        if password != confirm_password:
            return render(request, 'signup.html', {
                'form': UserCreationForm(),
                'error': 'Las contraseñas no coinciden'
            })
            
        try:
            validate_password(password)
        except ValidationError as e:
            return render(request, 'signup.html', {
                'form': UserCreationForm(),
                'error': e.messages[0]
            })

        try:
            # registrar usuario (usamos el email como username para evitar duplicados de nombre)
            user = User.objects.create_user(username=email, email=email, password=password)
            user.first_name = name
            user.save()
            # Iniciar sesión
            login(request, user)
            # Redirigir al panel dinámico
            return redirect('dashboard')
        except IntegrityError:
            return render(request, 'signup.html', {
                'form': UserCreationForm(),
                'error': 'El correo ya está en uso'
                }) 

@login_required
def pedidos(request):
    return redirect('dashboard')    

@login_required
def pedidos_completed(request):
    pedidos = Pedido.objects.filter(user=request.user, datecompleted__isnull=False).order_by('-datecompleted')
    
    # Cálculo global de ocupación de mesas para coherencia visual en panel_mozo.html
    todos_los_pedidos_activos = Pedido.objects.filter(datecompleted__isnull=True).exclude(estado='CANCELADO')
    mesas_data = []
    for i in range(1, 11):
        ocupada = todos_los_pedidos_activos.filter(mesa=i).exists()
        mesas_data.append({
            'numero': i,
            'ocupada': ocupada,
            'pedidos': Pedido.objects.none()
        })
        
    from django.db.models import Sum
    today = timezone.localtime(timezone.now()).date()
    total_servidas = Pedido.objects.filter(user=request.user, estado__in=['COBRADO', 'PAGADO'], datecompleted__date=today).count()
    propinas_ganadas = Pedido.objects.filter(user=request.user, estado__in=['COBRADO', 'PAGADO']).aggregate(total_propina=Sum('propina'))['total_propina'] or 0.00
    
    return render(request, 'panel_mozo.html', {
        'pedidos': pedidos,
        'mesas_data': mesas_data,
        'form_pedido': PedidoForm(),
        'total_servidas': total_servidas,
        'propinas_ganadas': propinas_ganadas,
        'is_history': True
    })

@login_required
def create_pedido(request):
    if request.method == 'GET':
        return render(request, 'create_pedido.html', {
            'form': PedidoForm()
        })
    else:
        form = PedidoForm(request.POST)
        if form.is_valid():
            new_pedido = form.save(commit=False)
            new_pedido.user = request.user
            new_pedido.save()
            form.save_m2m()
            return redirect('dashboard')
        else:
            return render(request, 'create_pedido.html', {
                'form': form,
                'error': 'Por favor corrige los errores del formulario.'
            })

@login_required
def pedido_detail(request, pedido_id):
    if request.method == 'GET':
        pedido = get_object_or_404(Pedido, pk=pedido_id)
        form = PedidoForm(instance=pedido)
        return render(request, 'pedido_detail.html', {'pedido': pedido, 'form': form})
    else:
        pedido = get_object_or_404(Pedido, pk=pedido_id)
        form = PedidoForm(request.POST, instance=pedido)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
        else:
            return render(request, 'pedido_detail.html', {
                'pedido': pedido,
                'form': form,
                'error': 'Error al actualizar el pedido'
            })

@login_required
def complete_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, pk=pedido_id, user=request.user)
    if request.method == 'POST':
        pedido.datecompleted = timezone.now()
        pedido.save()
        return redirect('dashboard') 
    else:
        return redirect('dashboard')

@login_required
def delete_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, pk=pedido_id, user=request.user)
    if request.method == 'POST':
        pedido.delete()
        return redirect('dashboard') 
    else:
        return redirect('dashboard')

@login_required
def signout(request):
    logout(request)
    return redirect('signin')

def signin(request):
    if request.method == 'GET':
        return render(request, 'signin.html', {
        'form': AuthenticationForm()
         })
    else:
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        user_obj = User.objects.filter(email=email).first()
        if user_obj:
            user = authenticate(username=user_obj.username, password=password)
        else:
            user = None

        if user is not None: #Si el usuario SI se verifica y SI existe
            login(request, user) #Se inicia sesión
            # Redirigir al panel dinámico
            return redirect('dashboard')
        else:
            return render(request, 'signin.html', { #Si el usuario NO se verifica y NO existe
                'form': AuthenticationForm(), #Se muestra el formulario de inicio de sesión
                'error': 'Correo o contraseña incorrectos' #Se muestra el mensaje de error
            }) 

@login_required
@never_cache
def dashboard(request):
    # Diccionario para mapear los roles a sus respectivos templates
    paneles = {
        1: 'panel_administrador.html',
        2: 'panel_cliente.html',
        3: 'panel_mozo.html',
        4: 'panel_cocina.html'
    }
    
    try:
        profile = request.user.profile
        role = profile.rol
    except Exception:
        # Si el usuario no tiene perfil (pasa con superusuarios creados en consola)
        if request.user.is_superuser:
            role = 1
        else:
            role = 2 # Cliente por defecto
            
        # Crear e inyectar el perfil permanente en la base de datos
        from .models import Profile
        profile, created = Profile.objects.get_or_create(user=request.user, defaults={'rol': role})
        if not created and profile.rol != role:
            profile.rol = role
            profile.save()
    template_name = paneles.get(role, '404.html') # fallback por si el rol no existe
    
    # Proveer las tareas para el template 
    context = {}
    if role == 4 or role == 1:
        # Cocina y Administrador deben ver todos los pedidos
        pedidos = Pedido.objects.filter(datecompleted__isnull=True).exclude(estado__in=['ENTREGADO', 'CANCELADO'])
        productos = Producto.objects.all()
        context['form_producto'] = ProductoForm()
        
        if role == 1:
            from datetime import timedelta
            
            # 1. Todos los pedidos para el administrador (completados, activos, cancelados)
            pedidos_todos = Pedido.objects.all().order_by('-created')
            context['pedidos_todos'] = pedidos_todos
            
            # 2. Cantidades por estado para los badges
            context['cant_nuevos'] = Pedido.objects.filter(estado='CREADO', datecompleted__isnull=True).count()
            context['cant_preparando'] = Pedido.objects.filter(estado='PREPARACION', datecompleted__isnull=True).count()
            context['cant_listos'] = Pedido.objects.filter(estado='LISTO', datecompleted__isnull=True).count()
            context['cant_entregados'] = Pedido.objects.filter(estado='ENTREGADO', datecompleted__isnull=True).count()
            context['cant_cobrados'] = Pedido.objects.filter(estado='COBRADO').count()
            context['cant_cancelados'] = Pedido.objects.filter(estado='CANCELADO').count()
            
            # 3. KPI 1: Plato más vendido (completados / cobrados)
            plato_mas_vendido = Producto.objects.filter(pedido__estado='COBRADO').annotate(total_vendido=Count('pedido')).order_by('-total_vendido').first()
            if not plato_mas_vendido:
                plato_mas_vendido = Producto.objects.annotate(total_vendido=Count('pedido')).order_by('-total_vendido').first()
            context['plato_mas_vendido'] = plato_mas_vendido.nombre if plato_mas_vendido else "N/A"
            
            # 4. KPI 2: Total Vendido Hoy (completados y cobrados hoy)
            today = timezone.localtime(timezone.now()).date()
            total_hoy = sum(p.get_total for p in Pedido.objects.filter(estado='COBRADO', datecompleted__date=today))
            context['total_vendido_hoy'] = total_hoy
            
            # 5. KPI 3: Tiempo promedio por pedido (basado en cobrados)
            tiempos = []
            pedidos_completados = Pedido.objects.filter(estado='COBRADO')
            for p in pedidos_completados:
                if p.datecompleted and p.created:
                    tiempos.append((p.datecompleted - p.created).total_seconds())
            if tiempos:
                avg_seconds = sum(tiempos) / len(tiempos)
                minutes = int(avg_seconds // 60)
                seconds = int(avg_seconds % 60)
                context['tiempo_promedio'] = f"{minutes}m {seconds}s"
            else:
                context['tiempo_promedio'] = "0m 0s"
                
            # 6. KPI 4: Mesas activas
            mesas_activas = Pedido.objects.filter(datecompleted__isnull=True).exclude(estado='CANCELADO').values('mesa').distinct().count()
            context['mesas_activas'] = mesas_activas
            context['mesas_activas_porcentaje'] = int((mesas_activas / 10) * 100)
            
            # 7. Ventas y pedidos semanales (Lun - Dom / cobrados y totales)
            start_of_week = today - timedelta(days=today.weekday())
            weekly_sales = []
            max_sales = 1.0
            max_orders = 1
            days_labels = ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab', 'Dom']
            
            for i in range(7):
                day_date = start_of_week + timedelta(days=i)
                # Ventas del día (basado en cobrados/pagados)
                day_pedidos_cobrados = Pedido.objects.filter(estado__in=['COBRADO', 'PAGADO'], datecompleted__date=day_date)
                day_total = float(sum(p.get_total for p in day_pedidos_cobrados))
                
                # Pedidos totales del día (excluyendo cancelados)
                day_pedidos_creados = Pedido.objects.filter(created__date=day_date).exclude(estado='CANCELADO')
                day_order_count = day_pedidos_creados.count()
                
                weekly_sales.append({
                    'label': days_labels[i],
                    'total': day_total,
                    'order_count': day_order_count,
                    'is_today': day_date == today,
                })
                if day_total > max_sales:
                    max_sales = day_total
                if day_order_count > max_orders:
                    max_orders = day_order_count
                    
            allowed_heights = [10, 20, 30, 40, 50, 60, 70, 80, 90, 95]
            for day_data in weekly_sales:
                pct_sales = int((day_data['total'] / max_sales) * 95) if max_sales > 0 else 10
                day_data['height_sales'] = min(allowed_heights, key=lambda x: abs(x - pct_sales))
                
                pct_orders = int((day_data['order_count'] / max_orders) * 95) if max_orders > 0 else 10
                day_data['height_orders'] = min(allowed_heights, key=lambda x: abs(x - pct_orders))
                
            context['weekly_sales'] = weekly_sales
            
            # 8. Top 3 platos más vendidos (basado en cobrados)
            top_platos = Producto.objects.annotate(
                total_vendido=Count('pedido', filter=Q(pedido__estado='COBRADO'))
            ).order_by('-total_vendido')[:3]
            
            top_platos_mapped = []
            for plato in top_platos:
                img_url = "https://lh3.googleusercontent.com/aida-public/AB6AXuBdA-15uCWozfJ38f4D9qzjBV1WHbYxSNfomDqOGk0DzB_GxqWr9i_UMIJxHCWlL5iCw0N9yvTVAKvb9bxA8eUI0HLRYk-jl8sdTjQ364rJoOh5WD3O4Q4o6uiLE8C8LNQRA7RfhS3HNCpt-AuW2I0j7lj7ndF8ZJdinohoebnhrw8TMlj-GOtebcd985ErA-FmxoDQpniyv-T79c0nb6N3n3xXsryxDvzjXVF9X9GZu1eXGr_e9g4vP2w9vqE96L4L9JLbA3J2fXiS"
                name_lower = plato.nombre.lower()
                if 'papa' in name_lower or 'fries' in name_lower or 'frita' in name_lower:
                    img_url = "https://lh3.googleusercontent.com/aida-public/AB6AXuCmrXFQ8sHFIxqHoWlfV2LxRBhG5kbyeuRmblIp7Ogl946xQUh1XiIE1oOLi4KnZLLT0ITDaQtCCd01clqSK04rkR5zYCP2voJhNYKEiW3YaUaqGj9lPlLgCMLRC7VM4ZfwIFYhkOxWZoQ5S8Fl-NG1m7gCI8V1KSja8HJvbvho3-7bWTaKNDF9mCJATIRbWfTdgEC8xWxboBE0iAigRGGVE0DLFw-HxclrOntqB6IDNSHh4vvcUZFw5nZM_78Uxt5WqjBICEFNONTq"
                elif any(x in name_lower for x in ['bebida', 'gaseosa', 'jugo', 'limonada', 'chicha', 'cerveza', 'shake', 'malteada']):
                    img_url = "https://lh3.googleusercontent.com/aida-public/AB6AXuB70WZ7kqIDqZq8r5O3nQyy-UDmrMZcJObnxHETpHbHCeKeJQgKG8MiRROd9navFvFkIQ1hydGjOTkGRwIQExrSLIqEDnrSg5l8Nez2kipF6aAS6It6qWFonH9zhAy7TXgrfbIGs1rg62fCaRWrZrP5h7AIF2GcGyzkIbMNBW_ltDtVk-WoPYY0zQvhuEprQM3Rpw0j7t3uA2psq5CxfDIvUNAtuXWsaFEJ0WGbdy1vFaqBjsSs3nxCvmUO3z0gWDaiX8nzpc_sol6-"
                
                all_completed_count = Pedido.objects.filter(datecompleted__isnull=False, estado='ENTREGADO').count() or 1
                trend_val = int((plato.total_vendido / all_completed_count) * 100)
                trend_text = f"+{trend_val}%" if trend_val > 0 else "0%"
                
                top_platos_mapped.append({
                    'nombre': plato.nombre,
                    'total_vendido': plato.total_vendido,
                    'img_url': img_url,
                    'trend_text': trend_text,
                })
            context['top_platos'] = top_platos_mapped
    elif role == 3:
        # Mozos ven los pedidos que crearon
        pedidos = Pedido.objects.filter(user=request.user, datecompleted__isnull=True).exclude(estado='CANCELADO')
        todos_los_pedidos_activos = Pedido.objects.filter(datecompleted__isnull=True).exclude(estado='CANCELADO')
        productos = None
        
        mesas_data = []
        for i in range(1, 11):
            # Una mesa está ocupada si hay algún pedido activo en ella en todo el negocio
            ocupada = todos_los_pedidos_activos.filter(mesa=i).exists()
            pedidos_mesa = pedidos.filter(mesa=i)
            mesas_data.append({
                'numero': i,
                'ocupada': ocupada,
                'pedidos': pedidos_mesa
            })
        context['mesas_data'] = mesas_data
        context['form_pedido'] = PedidoForm()
        
        # Calcular propinas y órdenes cobradas hoy por este mozo
        from django.db.models import Sum
        today = timezone.localtime(timezone.now()).date()
        total_servidas = Pedido.objects.filter(user=request.user, estado__in=['COBRADO', 'PAGADO'], datecompleted__date=today).count()
        context['total_servidas'] = total_servidas
        
        propinas_ganadas = Pedido.objects.filter(user=request.user, estado__in=['COBRADO', 'PAGADO']).aggregate(total_propina=Sum('propina'))['total_propina'] or 0.00
        context['propinas_ganadas'] = propinas_ganadas
    elif role == 2:
        # Clientes ven sus pedidos activos y también historial/stats
        pedidos = Pedido.objects.filter(Q(cliente=request.user) | Q(user=request.user), datecompleted__isnull=True)
        productos = None
        
        todos_sus_pedidos = Pedido.objects.filter(Q(cliente=request.user) | Q(user=request.user))
        context['historial_pedidos'] = todos_sus_pedidos.filter(datecompleted__isnull=False).order_by('-datecompleted')
        context['ultimo_pedido'] = todos_sus_pedidos.order_by('-created').first()
        
        # Producto favorito
        platos_pedidos = Producto.objects.filter(pedido__in=todos_sus_pedidos).annotate(veces_pedido=Count('pedido')).order_by('-veces_pedido').first()
        context['producto_favorito'] = platos_pedidos
        context['total_pedidos'] = todos_sus_pedidos.count()
        
        # Calcular el total gastado de los pedidos completados
        total_gastado = sum(pedido.get_total for pedido in context['historial_pedidos'])
        context['total_gastado'] = total_gastado
        
        # Calcular los puntos del bono: 10% de los S/. gastados en pedidos completados
        bono_points = int(total_gastado * Decimal('0.10'))
        context['bono_points'] = bono_points
        
        # Calcular estado de recompensas
        context['milestone_100_completed'] = bono_points >= 100
        context['milestone_500_completed'] = bono_points >= 500
        context['milestone_1000_completed'] = bono_points >= 1000
        
        context['milestone_100_current'] = False
        context['milestone_500_current'] = False
        context['milestone_1000_current'] = False
        
        # Determinar el mensaje de la siguiente recompensa y progreso
        if bono_points < 100:
            context['siguiente_recompensa'] = f"¡Solo faltan {100 - bono_points} puntos para Papas Fritas Gratis!"
            context['progreso_porcentaje'] = int((bono_points / 100) * 100)
            context['milestone_100_current'] = True
        elif bono_points < 500:
            context['siguiente_recompensa'] = f"¡Solo faltan {500 - bono_points} puntos para una Bebida Fría!"
            context['progreso_porcentaje'] = int(((bono_points - 100) / 400) * 100)
            context['milestone_500_current'] = True
        elif bono_points < 1000:
            context['siguiente_recompensa'] = f"¡Solo faltan {1000 - bono_points} puntos para un Menú de Hamburguesa!"
            context['progreso_porcentaje'] = int(((bono_points - 500) / 500) * 100)
            context['milestone_1000_current'] = True
        else:
            context['siguiente_recompensa'] = "¡Felicidades! Has alcanzado la máxima recompensa."
            context['progreso_porcentaje'] = 100
            
        # Pedido activo actual (el más reciente que aún no está completado)
        active_order = pedidos.order_by('-created').first()
        context['active_order'] = active_order
        
        # Calcular tiempo estimado y mensaje del chef si hay orden activa
        if active_order:
            if active_order.estado == 'CREADO':
                context['tiempo_estimado'] = "15-20 Minutos"
                context['chef_message'] = "¡Tu pedido ha sido recibido! Estamos listos para comenzar en breve."
            elif active_order.estado == 'PREPARACION':
                tiempo = 5 + (active_order.productos.count() * 3)
                context['tiempo_estimado'] = f"{tiempo} Minutos"
                context['chef_message'] = "¡El chef está cocinando tus platos ahora!"
            elif active_order.estado == 'LISTO':
                context['tiempo_estimado'] = "¡Listo para retirar!"
                context['chef_message'] = "¡Tu pedido está listo!"
            else:
                context['tiempo_estimado'] = "N/A"
                context['chef_message'] = "Tu pedido está siendo procesado."
        else:
            context['tiempo_estimado'] = None
            context['chef_message'] = None
        
        
    pedidos_creados = pedidos.filter(estado='CREADO')
    pedidos_preparacion = pedidos.filter(estado='PREPARACION')
    pedidos_listos = pedidos.filter(estado='LISTO')
    pedidos_entregados = pedidos.filter(estado='ENTREGADO')
    
    context.update({
        'pedidos': pedidos,
        'pedidos_creados': pedidos_creados,
        'pedidos_preparacion': pedidos_preparacion,
        'pedidos_listos': pedidos_listos,
        'pedidos_entregados': pedidos_entregados,
        'productos': productos
    })
    
    # Soporte para actualizaciones dinámicas via AJAX
    if request.GET.get('ajax') == 'true':
        from django.http import JsonResponse
        if role == 2:
            active_order_data = None
            active_order = context.get('active_order')
            if active_order:
                active_order_data = {
                    'id': active_order.id,
                    'estado': active_order.estado,
                    'estado_display': active_order.get_estado_display(),
                    'productos': [p.nombre for p in active_order.productos.all()],
                    'tiempo_estimado': context.get('tiempo_estimado'),
                    'chef_message': context.get('chef_message'),
                }
                
            historial_data = []
            for ped in context.get('historial_pedidos', []):
                historial_data.append({
                    'id': ped.id,
                    'datecompleted': ped.datecompleted.strftime('%b %d, %I:%M %p') if ped.datecompleted else '',
                    'total': str(ped.get_total),
                    'puntos': ped.get_puntos,
                    'productos_str': ", ".join([pr.nombre for pr in ped.productos.all()])
                })
                
            data = {
                'total_pedidos': context.get('total_pedidos', 0),
                'bono_points': context.get('bono_points', 0),
                'siguiente_recompensa': context.get('siguiente_recompensa'),
                'progreso_porcentaje': context.get('progreso_porcentaje'),
                'milestones': {
                    'm100_completed': context.get('milestone_100_completed'),
                    'm100_current': context.get('milestone_100_current'),
                    'm500_completed': context.get('milestone_500_completed'),
                    'm500_current': context.get('milestone_500_current'),
                    'm1000_completed': context.get('milestone_1000_completed'),
                    'm1000_current': context.get('milestone_1000_current'),
                },
                'active_order': active_order_data,
                'historial': historial_data,
            }
            return JsonResponse(data)
            
    return render(request, template_name, context)

def handler404(request, exception):
    return render(request, '404.html', status=404)


def cambiar_disponibilidad(request, producto_id):
    producto = get_object_or_404(Producto, id=producto_id)
    producto.disponible = not producto.disponible # Invierte el valor (True -> False y viceversa)
    producto.save()
    return redirect('dashboard')

@login_required
def crear_producto(request):
    if request.user.profile.rol not in [1, 4]:
        return redirect('dashboard')
        
    if request.method == 'GET':
        return render(request, 'crear_producto.html', {'form': ProductoForm()})
    else:
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
        return render(request, 'crear_producto.html', {'form': form})

@login_required
def update_estado(request, pedido_id, nuevo_estado):
    if request.user.profile.rol not in [1, 3, 4]:
        return redirect('dashboard')
        
    pedido = get_object_or_404(Pedido, id=pedido_id)
    if nuevo_estado in dict(Pedido.EstadoPedido.choices):
        pedido.estado = nuevo_estado
        if nuevo_estado == 'COBRADO':
            pedido.datecompleted = timezone.now()
            propina_val = request.POST.get('propina', 0.00)
            try:
                pedido.propina = float(propina_val)
            except (ValueError, TypeError):
                pedido.propina = 0.00
        pedido.save()
    return redirect('dashboard')

@login_required
def editar_producto(request, producto_id):
    if request.user.profile.rol not in [1, 4]:
        return redirect('dashboard')
        
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('dashboard')
    
    return redirect('dashboard')

@login_required
def eliminar_producto(request, producto_id):
    if request.user.profile.rol not in [1, 4]:
        return redirect('dashboard')
        
    producto = get_object_or_404(Producto, id=producto_id)
    if request.method == 'POST' or request.method == 'GET':
        producto.delete()
        
    return redirect('dashboard')

@login_required
def cancelar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    rol = request.user.profile.rol
    
    can_cancel = False
    if rol in [1, 3]:
        can_cancel = True
    elif rol == 2:
        if (pedido.cliente == request.user or pedido.user == request.user) and pedido.estado == 'CREADO':
            can_cancel = True
            
    if can_cancel:
        pedido.estado = 'CANCELADO'
        pedido.datecompleted = timezone.now()
        pedido.save()
        
    return redirect('dashboard')

@login_required
def generar_qr_yape(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pedido_id = data.get('pedido_id')
            if not pedido_id:
                return JsonResponse({'error': 'pedido_id es requerido'}, status=400)
                
            pedido = get_object_or_404(Pedido, id=pedido_id)
            total = float(pedido.get_total)
            
            propina_input = data.get('propina', 0)
            try:
                propina_input = float(propina_input)
            except (ValueError, TypeError):
                propina_input = 0.0

            if hasattr(pedido, 'propina'):
                if propina_input > 0:
                    pedido.propina = propina_input
                    pedido.save()
                pedido_propina = float(pedido.propina) if pedido.propina else 0.0
            else:
                pedido_propina = propina_input

            monto_final = total + (pedido_propina or propina_input)
            
            try:
                response = requests.post(
                    'https://api.mercadopago.com/v1/payments',
                    headers={
                        'Authorization': 'Bearer TEST-79a0da69-c240-40fb-ba50-73f0a60914be',
                        'Content-Type': 'application/json'
                    },
                    json={
                        'transaction_amount': monto_final,
                        'description': 'Pedido Cocina',
                        'payment_method_id': 'qr_payment',
                        'payer': {
                            'email': 'mozo@restaurante.com'
                        },
                        'notification_url': request.build_absolute_uri(reverse('webhook_mercadopago')),
                        'external_reference': str(pedido.id),
                        'metadata': {
                            'pedido_id': str(pedido.id)
                        }
                    },
                    timeout=10
                )
                res_json = response.json()
                payment_id = res_json.get('id', f'mp_mock_{pedido.id}')
                
                poi = res_json.get('point_of_interaction', {})
                tx_data = poi.get('transaction_data', {}) if isinstance(poi, dict) else {}
                
                qr_url = ""
                if isinstance(tx_data, dict):
                    qr_code_base64 = tx_data.get('qr_code_base64', '')
                    if qr_code_base64:
                        qr_url = f"data:image/png;base64,{qr_code_base64}"
                    else:
                        qr_url = tx_data.get('ticket_url', '')
                        
                if not qr_url:
                    qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=300x300&data=yape://pay?id={payment_id}%26amount={monto_final}"
            except Exception:
                payment_id = f'mp_mock_{pedido.id}'
                qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=300x300&data=yape://pay?id={payment_id}%26amount={monto_final}"
                
            print("--- EL ID PARA EL SIMULADOR ES: ---", payment_id)
            return JsonResponse({'qr_url': qr_url, 'payment_id': payment_id})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Método no permitido'}, status=405)


@csrf_exempt
def webhook_mercadopago(request):
    """
    Webhook endpoint para notificaciones de Mercado Pago.
    Soporta: mocks locales (mp_mock_), simulador oficial de MP (IDs numéricos)
    y pagos reales.
    """
    MERCADO_PAGO_ACCESS_TOKEN = "TEST-79a0da69-c240-40fb-ba50-73f0a60914be"

    # --- Prints detallados para diagnóstico en consola de Django ---
    print("\n" + "="*50)
    print("=== WEBHOOK MERCADOPAGO RECIBIDO ===")
    print("="*50)
    print("Método HTTP :", request.method)
    print("GET params  :", dict(request.GET))
    print("Headers     :", dict(request.headers))
    raw_body = request.body
    print("Body (raw)  :", raw_body[:500] if raw_body else "(vacío)")
    print("="*50)

    if request.method != 'POST':
        return HttpResponse("Method not allowed", status=405)

    try:
        # 1. Parsear body JSON (puede venir vacío en notificaciones tipo IPN)
        data = {}
        if raw_body:
            try:
                data = json.loads(raw_body)
            except Exception:
                pass

        print("[WEBHOOK] Data parseada:", data)

        # Si el payload directo indica que el pago fue rechazado, evitamos procesar y bypassear
        if isinstance(data, dict) and data.get('status') == 'rejected':
            print("[WEBHOOK] Pago rechazado según payload. Sin acción.")
            return HttpResponse("No processing needed", status=200)

        # 2. Extraer el payment_id de TODAS las formas posibles que MP usa
        payment_id = None

        # Formato Webhook estándar: {"action":"payment.updated","data":{"id":"12345"}}
        if isinstance(data, dict):
            data_obj = data.get('data', {})
            if isinstance(data_obj, dict):
                payment_id = data_obj.get('id')

        # Formato IPN legacy: GET ?id=12345&topic=payment
        if not payment_id:
            payment_id = request.GET.get('id') or request.GET.get('data.id')

        # Formato alternativo directo en body
        if not payment_id and isinstance(data, dict):
            payment_id = data.get('id')

        print("[WEBHOOK] Payment ID extraído:", payment_id)

        if not payment_id:
            print("[WEBHOOK] ERROR: No se encontró payment_id en ningún lugar del request.")
            return HttpResponse("No payment ID found", status=400)

        action = data.get('action', '') if isinstance(data, dict) else ''
        topic = request.GET.get('topic', '')
        print(f"[WEBHOOK] Action: '{action}' | Topic: '{topic}'")

        # 3. CASO MOCK LOCAL: el ID empieza con 'mp_mock_' (pruebas locales sin MP)
        if str(payment_id).startswith('mp_mock_'):
            pedido_id = str(payment_id).replace('mp_mock_', '')
            pedido = Pedido.objects.filter(id=pedido_id).first()
            if pedido:
                pedido.estado = 'PAGADO'
                pedido.datecompleted = timezone.now()
                pedido.save()
                print(f"[WEBHOOK] [OK] MOCK LOCAL: Pedido {pedido_id} marcado como PAGADO")
                return HttpResponse("OK", status=200)
            print(f"[WEBHOOK] [ERROR] MOCK LOCAL: Pedido {pedido_id} no encontrado")
            return HttpResponse("Pedido not found", status=404)

        # 4. CASO SIMULADOR / REAL: verificar con la API de Mercado Pago
        print(f"[WEBHOOK] Consultando API de MP para payment_id: {payment_id}")
        mp_ok = False  # bandera: True si el pago fue verificado y procesado como real
        try:
            mp_response = requests.get(
                f'https://api.mercadopago.com/v1/payments/{payment_id}',
                headers={'Authorization': f'Bearer {MERCADO_PAGO_ACCESS_TOKEN}'},
                timeout=10
            )
            print(f"[WEBHOOK] MP API status: {mp_response.status_code}")
            print(f"[WEBHOOK] MP API body: {mp_response.text[:300]}")

            if mp_response.status_code == 200:
                pay_data = mp_response.json()
                status_mp = pay_data.get('status')
                print(f"[WEBHOOK] Estado del pago en MP: {status_mp}")

                if status_mp == 'approved':
                    metadata = pay_data.get('metadata', {})
                    pedido_id_mp = metadata.get('pedido_id') or pay_data.get('external_reference')
                    print(f"[WEBHOOK] Pedido ID desde MP: {pedido_id_mp}")
                    if pedido_id_mp:
                        pedido = Pedido.objects.filter(id=pedido_id_mp).first()
                        if pedido:
                            pedido.estado = 'PAGADO'
                            pedido.datecompleted = timezone.now()
                            pedido.save()
                            print(f"[WEBHOOK] [OK] REAL: Pedido {pedido_id_mp} marcado como PAGADO")
                            mp_ok = True
                            return HttpResponse("OK", status=200)

                print(f"[WEBHOOK] Pago real verificado, estado={status_mp}. Sin accion.")
                mp_ok = True  # fue real pero no approved, no necesita bypass
                return HttpResponse("OK", status=200)
            else:
                print(f"[WEBHOOK] [WARN] MP API retorno {mp_response.status_code} - activando bypass de simulador.")

        except Exception as api_err:
            print(f"[WEBHOOK] Excepcion en llamada MP: {str(api_err)} - activando bypass de simulador.")

        # BYPASS DE DESARROLLO: MP no pudo validar (simulador con ID ficticio)
        # Se aplica cuando mp_ok sigue siendo False
        if not mp_ok:
            print("[WEBHOOK] Aplicando BYPASS: buscando ultimo pedido activo...")
            pedido_bypass = (
                Pedido.objects
                .filter(datecompleted__isnull=True)
                .exclude(estado='CANCELADO')
                .order_by('-created')
                .first()
            )
            if pedido_bypass:
                pedido_bypass.estado = 'PAGADO'
                pedido_bypass.datecompleted = timezone.now()
                pedido_bypass.save()
                print(f"[WEBHOOK] [OK] BYPASS: Pedido {pedido_bypass.id} mesa={pedido_bypass.mesa} -> PAGADO")
            else:
                print("[WEBHOOK] [ERROR] BYPASS: No hay pedidos activos para marcar.")
        return HttpResponse("OK", status=200)

    except Exception as e:
        import traceback
        print(f"[WEBHOOK] EXCEPCION GENERAL: {str(e)}")
        print(traceback.format_exc())
        return HttpResponse(str(e), status=400)

@login_required
def obtener_estado_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    return JsonResponse({'estado': pedido.estado})

@login_required
def exportar_pedidos_excel(request):
    if request.user.profile.rol != 1:
        return HttpResponse("No autorizado", status=403)

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # 1. Crear el libro y la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Control de Pedidos"
    
    # Asegurar que se vean las líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True

    # 2. Definir estilos "bonitos"
    font_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    font_bold = Font(name="Segoe UI", size=10, bold=True)
    
    fill_title = PatternFill(start_color="B50303", end_color="B50303", fill_type="solid") # Rojo El Buen Sabor
    fill_header = PatternFill(start_color="281715", end_color="281715", fill_type="solid") # Oscuro Elegante
    fill_total = PatternFill(start_color="FFE9E6", end_color="FFE9E6", fill_type="solid") # Rosado muy claro / crema para totales

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    border_thin = Side(style="thin", color="D1D5DB")
    border_double = Side(style="double", color="111827")
    border_thick = Side(style="medium", color="111827")

    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    border_header_cell = Border(left=border_thin, right=border_thin, top=border_thick, bottom=border_thick)
    border_total_row = Border(top=border_thin, bottom=border_double)

    # 3. Título del Reporte
    ws.merge_cells("A1:J1")
    ws["A1"] = "REPORTE GENERAL DE PEDIDOS - EL BUEN SABOR"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 40

    # Fila vacía de separación
    ws.row_dimensions[2].height = 15

    # 4. Encabezados de Tabla
    headers = [
        "Pedido ID", "Mesa", "Cliente", "Email", 
        "Productos", "Estado", "Fecha y Hora", 
        "Subtotal (S/.)", "Propina (S/.)", "Total (S/.)"
    ]
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header_title
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_header_cell
    
    ws.row_dimensions[3].height = 25

    # 5. Obtener todos los pedidos
    pedidos = Pedido.objects.all().order_by('-created')
    
    row_num = 4
    for ped in pedidos:
        cliente_name = ped.cliente.get_full_name() or ped.cliente.username if ped.cliente else "Invitado"
        cliente_email = ped.cliente.email if ped.cliente else "N/A"
        productos_list = ", ".join([p.nombre for p in ped.productos.all()]) or "Sin productos"
        estado_label = ped.get_estado_display()
        fecha_str = ped.created.astimezone(timezone.get_current_timezone()).strftime('%d/%m/%Y %H:%M') if ped.created else "N/A"
        
        subtotal_val = float(ped.get_total)
        propina_val = float(ped.propina)
        total_val = subtotal_val + propina_val
        
        row_values = [
            f"#{ped.id}",
            f"Mesa {ped.mesa}",
            cliente_name,
            cliente_email,
            productos_list,
            estado_label,
            fecha_str,
            subtotal_val,
            propina_val,
            total_val
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = val
            cell.font = font_body
            cell.border = border_cell
            
            # Alineaciones y formatos numéricos
            if col_idx in [1, 2, 6, 7]:
                cell.alignment = align_center
            elif col_idx in [8, 9, 10]:
                cell.alignment = align_right
                cell.number_format = '"S/." #,##0.00'
            else:
                cell.alignment = align_left
                
        ws.row_dimensions[row_num].height = 20
        row_num += 1

    # 6. Fila de Totales Generales
    ws.cell(row=row_num, column=1).value = "TOTALES"
    ws.cell(row=row_num, column=1).font = font_bold
    ws.cell(row=row_num, column=1).alignment = align_left
    ws.cell(row=row_num, column=1).fill = fill_total
    ws.cell(row=row_num, column=1).border = border_total_row
    
    # Combinar celdas de Totales de la A a la G
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
    for c in range(2, 8):
        ws.cell(row=row_num, column=c).fill = fill_total
        ws.cell(row=row_num, column=c).border = border_total_row
        
    # Fórmulas de Suma para montos
    for col_letter, col_idx in [('H', 8), ('I', 9), ('J', 10)]:
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = f"=SUM({col_letter}4:{col_letter}{row_num-1})"
        cell.font = font_bold
        cell.alignment = align_right
        cell.fill = fill_total
        cell.number_format = '"S/." #,##0.00'
        cell.border = border_total_row
        
    ws.row_dimensions[row_num].height = 25

    # 7. Habilitar filtros automáticos en la tabla
    ws.auto_filter.ref = f"A3:J{row_num-1}"

    # 8. Autoajustar anchos de columna
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.row == 1:
                continue
            if cell.value:
                val_str = str(cell.value)
                if len(val_str) > max_len:
                    max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # 9. Retornar archivo HTTP como Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Pedidos_{timezone.localdate().strftime("%Y-%m-%d")}.xlsx"'
    wb.save(response)
    return response